# from wsgiref.simple_server import make_server
# from pyramid.config import Configurator
# from pyramid.response import Response
# import os

# def hello_world(request):
#     name = os.environ.get('NAME')
#     if name == None or len(name) == 0:
#         name = "world"
#     message = "Hello, " + name + "!\n"
#     return Response(message)

# if __name__ == '__main__':
#     port = int(os.environ.get("PORT"))
#     with Configurator() as config:
#         config.add_route('hello', '/')
#         config.add_view(hello_world, route_name='hello')
#         app = config.make_wsgi_app()
#     server = make_server('0.0.0.0', port, app)
#     server.serve_forever()

from wsgiref.simple_server import make_server
from pyramid.config import Configurator
from pyramid.response import Response
from werkzeug.wsgi import DispatcherMiddleware
from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for, flash
from datetime import datetime, timedelta
from functools import wraps
import pandas as pd
import qrcode
import io
import base64
import sqlite3
import os

# Pyramid app for hello world
def hello_world(request):
    name = os.environ.get('NAME')
    if name == None or len(name) == 0:
        name = "world"
    message = "Hello, " + name + "!\n"
    return Response(message)

# Flask app for attendance system
flask_app = Flask(__name__)
flask_app.secret_key = 'dancesport_at_osu_secretary_2025_2026'

# Database file path
DATABASE = 'attendance.db'

# Admin credentials (in production, store these in the database with hashed passwords)
ADMIN_CREDENTIALS = {
    'dancesport': 'secretary2526', 
    #'teacher': 'teacher123' - example login
}

def init_db():
    """Initialize the database with required tables"""
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    
    # Create users table (for future expansion)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT DEFAULT 'admin',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Create attendance records table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS attendance_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            first_name TEXT NOT NULL,
            last_name TEXT,
            student_id TEXT,
            event TEXT NOT NULL,
            feedback TEXT,
            timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Create events table for managing recurring events
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            event_name TEXT UNIQUE NOT NULL,
            description TEXT,
            start_date DATE,
            end_date DATE,
            recurring_days TEXT,  -- JSON string of days [0,1,2,3,4,5,6] for Mon-Sun
            is_active BOOLEAN DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Create QR codes table (optional - could regenerate as needed)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS qr_codes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            event_name TEXT UNIQUE NOT NULL,
            qr_image_base64 TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Create students table for better tracking
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            first_name TEXT NOT NULL,
            last_name TEXT,
            student_id TEXT UNIQUE,
            email TEXT,
            phone TEXT,
            dues_status TEXT DEFAULT 'unpaid',  -- 'paid', 'unpaid', 'exempt'
            dues_paid_date DATE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    conn.commit()
    conn.close()

def get_db_connection():
    """Get database connection with row factory"""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row  # This allows accessing columns by name
    return conn

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            flash('Please log in to access this page.')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@flask_app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if username in ADMIN_CREDENTIALS and ADMIN_CREDENTIALS[username] == password:
            session['logged_in'] = True
            session['username'] = username
            flash(f'Welcome, {username}!')
            return redirect(url_for('home'))
        else:
            flash('Invalid username or password.')
    
    return render_template('login.html')

@flask_app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out.')
    return redirect(url_for('login'))

@flask_app.route('/admin')
@login_required
def home():
    return render_template('admin.html')

@flask_app.route('/generate_qr', methods=['POST'])
@login_required
def generate_qr():
    """Generate QR code for a specific event"""
    event_name = request.form.get('event_name')
    if not event_name:
        return "Event name is required", 400

    # URL for the attendance form
    form_url = request.url_root + f'form?event={event_name}'

    # Generate QR code
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(form_url)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")

    # Convert to base64 for display
    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    qr_image = base64.b64encode(buffer.getvalue()).decode()

    # Save QR code to database
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT OR REPLACE INTO qr_codes (event_name, qr_image_base64, updated_at)
        VALUES (?, ?, CURRENT_TIMESTAMP)
    ''', (event_name, qr_image))
    conn.commit()
    conn.close()

    return render_template('qr_display.html', qr_image=qr_image, event_name=event_name, form_url=form_url)

@flask_app.route('/form')
def attendance_form():
    """Show attendance form after scanning QR (no login required)"""
    event = request.args.get('event', '')
    return render_template('form.html', event=event)

@flask_app.route('/submit', methods=['POST'])
def submit():
    """Handle attendance form submission (no login required)"""
    first_name = request.form.get('first_name')
    last_name = request.form.get('last_name')
    student_id = request.form.get('student_id')
    event = request.form.get('event')
    feedback = request.form.get('feedback')
    
    # Insert into database
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute('''
        INSERT INTO attendance_records (first_name, last_name, student_id, event, feedback)
        VALUES (?, ?, ?, ?, ?)
    ''', (first_name, last_name, student_id, event, feedback))
    
    record_id = cursor.lastrowid
    
    # Get the inserted record for display
    cursor.execute('SELECT * FROM attendance_records WHERE id = ?', (record_id,))
    record = dict(cursor.fetchone())
    
    # Update or create student record
    if student_id:
        cursor.execute('''
            INSERT OR REPLACE INTO students (first_name, last_name, student_id, updated_at)
            VALUES (?, ?, ?, CURRENT_TIMESTAMP)
        ''', (first_name, last_name, student_id))
    
    conn.commit()
    conn.close()
    
    return render_template('success.html', record=record)

@flask_app.route('/view_data')
@login_required
def view_data():
    """View attendance data with QR codes"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Get all attendance records
    cursor.execute('''
        SELECT * FROM attendance_records 
        ORDER BY timestamp DESC
    ''')
    records = [dict(row) for row in cursor.fetchall()]
    
    # Get QR codes
    cursor.execute('SELECT event_name, qr_image_base64 FROM qr_codes')
    qr_data = cursor.fetchall()
    qr_codes = {row['event_name']: row['qr_image_base64'] for row in qr_data}
    
    conn.close()
    
    return render_template('data.html', records=records, qr_codes=qr_codes)

@flask_app.route('/attendance_tracker')
@login_required
def attendance_tracker():
    """New attendance tracking view"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Get all records for the tracker
    cursor.execute('''
        SELECT * FROM attendance_records 
        ORDER BY event, timestamp
    ''')
    records = [dict(row) for row in cursor.fetchall()]
    
    # Get unique events
    cursor.execute('SELECT DISTINCT event FROM attendance_records ORDER BY event')
    events = [row['event'] for row in cursor.fetchall()]
    
    conn.close()
    
    return render_template('attendance_tracker.html', records=records, events=events)

@flask_app.route('/generate_attendance_matrix', methods=['POST'])
@login_required
def generate_attendance_matrix():
    """Generate attendance matrix for a specific event/class"""
    try:
        data = request.get_json()
        event_name = data.get('event_name')
        start_date = datetime.strptime(data.get('start_date'), '%Y-%m-%d')
        end_date = datetime.strptime(data.get('end_date'), '%Y-%m-%d')
        
        if not event_name:
            return jsonify({'error': 'Event name is required'}), 400

        # Query database for this specific event
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT first_name, last_name, student_id, timestamp
            FROM attendance_records 
            WHERE LOWER(event) = LOWER(?)
            AND date(timestamp) BETWEEN ? AND ?
            ORDER BY first_name, last_name, timestamp
        ''', (event_name, start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))
        
        event_records = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        if not event_records:
            return jsonify({'error': 'No records found for this event in the specified date range'}), 400

        # Create student list
        students = {}
        for record in event_records:
            student_key = f"{record['first_name']} {record['last_name'] or ''}".strip()
            if student_key not in students:
                students[student_key] = {
                    'name': student_key,
                    'student_id': record.get('student_id', ''),
                    'attendances': {}
                }

        # Generate date columns for the specified range
        dates = []
        current_date = start_date
        while current_date <= end_date:
            dates.append(current_date.strftime('%Y-%m-%d'))
            current_date += timedelta(days=1)

        # Mark attendances
        for record in event_records:
            student_key = f"{record['first_name']} {record['last_name'] or ''}".strip()
            record_date = datetime.strptime(record['timestamp'], '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
            if record_date in dates:
                students[student_key]['attendances'][record_date] = 'X'

        # Create matrix data
        matrix_data = []
        for student_key, student_info in students.items():
            row = {
                'name': student_info['name'],
                'student_id': student_info['student_id'],
                'total': sum(1 for date in dates if student_info['attendances'].get(date) == 'X')
            }
            
            # Add attendance marks for each date
            for date in dates:
                row[date] = student_info['attendances'].get(date, '')
            
            matrix_data.append(row)

        # Sort by name
        matrix_data.sort(key=lambda x: x['name'])

        return jsonify({
            'success': True,
            'data': matrix_data,
            'dates': dates,
            'event_name': event_name
        })

    except Exception as error:
        print(f'Matrix generation error: {error}')
        return jsonify({'error': str(error)}), 500

@flask_app.route('/export_attendance_matrix', methods=['POST'])
@login_required
def export_attendance_matrix():
    """Export attendance matrix to Excel"""
    try:
        data = request.get_json()
        matrix_data = data.get('data', [])
        dates = data.get('dates', [])
        event_name = data.get('event_name', 'Event')
        
        if not matrix_data:
            return jsonify({'error': 'No data to export'}), 400

        # Create DataFrame
        df_data = []
        for row in matrix_data:
            excel_row = {
                'Student Name': row['name'],
                'Student ID': row['student_id'],
                'Total': row['total']
            }
            # Add date columns
            for date in dates:
                excel_row[date] = row.get(date, '')
            
            df_data.append(excel_row)

        df = pd.DataFrame(df_data)

        # Create Excel file
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=f'{event_name} Attendance', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets[f'{event_name} Attendance']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 20)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Style the header row
            from openpyxl.styles import Font, PatternFill, Alignment
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            center_align = Alignment(horizontal="center")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            
            # Center align the X marks and totals
            for row in worksheet.iter_rows(min_row=2):
                for cell in row[2:]:  # Skip name and ID columns
                    cell.alignment = center_align

        output.seek(0)
        
        filename = f"{event_name.replace(' ', '_')}_attendance_matrix.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as error:
        print(f'Excel export error: {error}')
        return jsonify({'error': f'Export failed: {str(error)}'}), 500

@flask_app.route('/api/attendance', methods=['GET'])
@login_required
def get_attendance_data():
    """API endpoint to get data as JSON"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM attendance_records ORDER BY timestamp DESC')
    records = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    return jsonify(records)

@flask_app.route('/export_to_excel', methods=['POST'])
@login_required
def export_to_excel():
    """Export data to Excel file (legacy function for compatibility)"""
    try:
        data = request.get_json()
        export_data = data.get('data', [])
        filename = data.get('filename', 'attendance_export.xlsx')

        if not export_data:
            return jsonify({'error': 'No data to export'}), 400

        # Convert data to pandas DataFrame
        if len(export_data) > 1:
            df = pd.DataFrame(export_data[1:], columns=export_data[0])
        else:
            return jsonify({'error': 'No data rows to export'}), 400

        # Create Excel file in memory
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Attendance Data', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Attendance Data']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Style the header row
            from openpyxl.styles import Font, PatternFill
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill

        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as error:
        print(f'Excel export error: {error}')
        return jsonify({'error': f'Export failed: {str(error)}'}), 500

# Database management routes
@flask_app.route('/manage_students')
@login_required
def manage_students():
    """Student management page"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT s.*, 
               COUNT(ar.id) as attendance_count,
               MAX(ar.timestamp) as last_attendance
        FROM students s
        LEFT JOIN attendance_records ar ON s.student_id = ar.student_id
        GROUP BY s.id
        ORDER BY s.first_name, s.last_name
    ''')
    students = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    return render_template('manage_students.html', students=students)

@flask_app.route('/update_dues_status', methods=['POST'])
@login_required
def update_dues_status():
    """Update student dues status"""
    data = request.get_json()
    student_id = data.get('student_id')
    dues_status = data.get('dues_status')
    
    if not student_id or dues_status not in ['paid', 'unpaid', 'exempt']:
        return jsonify({'error': 'Invalid data'}), 400
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    dues_paid_date = datetime.now().strftime('%Y-%m-%d') if dues_status == 'paid' else None
    
    cursor.execute('''
        UPDATE students 
        SET dues_status = ?, dues_paid_date = ?, updated_at = CURRENT_TIMESTAMP
        WHERE student_id = ?
    ''', (dues_status, dues_paid_date, student_id))
    
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})

if __name__ == '__main__':
    # Initialize database on startup
    init_db()
    
    # Get port from environment variable
    port = int(os.environ.get("PORT", 8080))
    
    # Create Pyramid app
    with Configurator() as config:
        config.add_route('hello', '/')
        config.add_view(hello_world, route_name='hello')
        pyramid_app = config.make_wsgi_app()
    
    # Combine Pyramid and Flask apps using DispatcherMiddleware
    # Flask app will handle all routes except '/' which goes to Pyramid
    application = DispatcherMiddleware(flask_app, {
        '/hello': pyramid_app
    })
    
    # Start the server
    server = make_server('0.0.0.0', port, application)
    print(f"Server running on http://0.0.0.0:{port}")
    print("Hello World (Pyramid): http://0.0.0.0:{}/hello".format(port))
    print("Attendance System (Flask): http://0.0.0.0:{}/admin".format(port))
    server.serve_forever()